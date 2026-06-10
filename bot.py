"""
ScoutBot — Railway Edition
"""

import os, cv2, csv, subprocess, threading, logging, asyncio, uuid, re
from pathlib import Path
from datetime import datetime

import numpy as np
import imageio_ffmpeg
FFMPEG = imageio_ffmpeg.get_ffmpeg_exe()
import yt_dlp
from flask import Flask, render_template, request, jsonify, send_file
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    CallbackQueryHandler, ConversationHandler, ContextTypes, filters
)

# ─── CONFIG ──────────────────────────────────────────────────────────────────
BOT_TOKEN  = os.environ.get("BOT_TOKEN", "8417374602:AAHgzLA5YJp3oEtCPklpgdYI-BqolIhDeW4")
CHAT_ID    = os.environ.get("CHAT_ID",   "7125492867")
PUBLIC_URL = os.environ.get("PUBLIC_URL") or "https://stats-production-a62b.up.railway.app"
ANALISE_URL = os.environ.get("ANALISE_URL", "https://web-production-c33c9.up.railway.app")
BOT_SECRET  = os.environ.get("BOT_SECRET", "scoutbot_secret_2024")
PORT       = int(os.environ.get("PORT", 5055))
WORK_DIR   = Path(os.environ.get("WORK_DIR", "/data"))
CSV_PATH   = WORK_DIR / "gols.csv"

SECONDS_BEFORE = 15
MIN_STATIC     = 2.0
DIFF_THRESHOLD = 25
PADDING_AFTER  = 2

WORK_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(format="%(asctime)s %(levelname)s %(name)s — %(message)s", level=logging.INFO)
log = logging.getLogger("scoutbot")
log.info(f"FFMPEG path: {FFMPEG}")

pending_clips: dict[str, dict] = {}
_frames_cache: dict[str, list] = {}
_frame_b64_cache: dict[str, str] = {}

# ─── SESSÕES EM DISCO ────────────────────────────────────────────────────────
def session_path(session_id: str) -> Path:
    return WORK_DIR / f"session_{session_id}.json"

def session_save(session_id: str, data: dict):
    import json
    to_save = {k: v for k, v in data.items() if k not in ("frame_b64", "frames")}
    with open(session_path(session_id), "w") as f:
        json.dump(to_save, f)

def session_load(session_id: str):
    import json
    if not session_id:
        return None
    p = session_path(session_id)
    if not p.exists():
        return None
    try:
        with open(p) as f:
            return json.load(f)
    except:
        return None

def session_delete(session_id: str):
    try: session_path(session_id).unlink()
    except: pass

# ─── PLANILHA EXCEL ──────────────────────────────────────────────────────────
XLSX_PATH = WORK_DIR / "scoutbot_dados.xlsx"

def xlsx_ensure():
    """Cria o Excel com abas formatadas se não existir."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if XLSX_PATH.exists():
            return

        wb = Workbook()

        # ── ABA GOLS ──────────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Gols"

        headers = ["Data", "Hora", "Competição", "Time Mandante", "Time Visitante",
                   "Time do Gol", "Timestamp Vídeo", "Nº Goleador", "Goleador",
                   "Nº Assistência", "Assistência", "Vídeo"]

        header_fill = PatternFill("solid", start_color="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        ws.row_dimensions[1].height = 30
        col_widths = [12, 8, 14, 18, 18, 14, 14, 10, 20, 12, 20, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

        # ── ABA RESUMO JOGADORES ──────────────────────────────────────────────
        ws2 = wb.create_sheet("Resumo por Jogador")
        h2 = ["Jogador", "Time", "Competição", "Gols", "Assistências", "Jogos"]
        for col, h in enumerate(h2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill("solid", start_color="1A5276")
            cell.alignment = header_align
            cell.border = border
        ws2.row_dimensions[1].height = 30
        for i, w in enumerate([22, 18, 14, 8, 12, 8], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = "A2"

        # ── ABA RESUMO TIMES ──────────────────────────────────────────────────
        ws3 = wb.create_sheet("Resumo por Time")
        h3 = ["Time", "Competição", "Jogos", "Gols Marcados", "Gols Sofridos"]
        for col, h in enumerate(h3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill("solid", start_color="145A32")
            cell.alignment = header_align
            cell.border = border
        ws3.row_dimensions[1].height = 30
        for i, w in enumerate([22, 14, 8, 14, 14], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.freeze_panes = "A2"

        wb.save(str(XLSX_PATH))
        log.info("✅ Planilha Excel criada")
    except Exception as e:
        log.error(f"xlsx_ensure error: {e}")

def xlsx_save(row: dict):
    """Adiciona uma linha na aba Gols e atualiza resumos."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
        from openpyxl.utils import get_column_letter

        xlsx_ensure()
        wb = load_workbook(str(XLSX_PATH))
        ws = wb["Gols"]

        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        even_fill = PatternFill("solid", start_color="EBF5FB")

        next_row = ws.max_row + 1
        values = [
            row.get("data", ""), row.get("hora", ""),
            row.get("competicao", ""), row.get("time_mand", ""),
            row.get("time_visit", ""), row.get("time_gol", ""),
            row.get("timestamp_video", ""), row.get("num_gol", ""),
            row.get("jogador_gol", ""), row.get("num_assist", ""),
            row.get("jogador_assist", ""), row.get("video_origem", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=next_row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if next_row % 2 == 0:
                cell.fill = even_fill

        # Atualiza aba Resumo por Jogador
        _atualizar_resumo_jogador(wb, row)
        _atualizar_resumo_time(wb, row)

        wb.save(str(XLSX_PATH))
    except Exception as e:
        log.error(f"xlsx_save error: {e}")

def _atualizar_resumo_jogador(wb, row):
    """Atualiza ou cria linha do jogador no resumo."""
    from openpyxl.styles import Alignment, Border, Side
    ws = wb["Resumo por Jogador"]
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    jogador = row.get("jogador_gol", "")
    time_gol = row.get("time_gol", "")
    comp = row.get("competicao", "")
    assist = row.get("jogador_assist", "")

    # Busca ou cria linha do goleador
    found = False
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == jogador and ws.cell(r, 3).value == comp:
            ws.cell(r, 4).value = (ws.cell(r, 4).value or 0) + 1
            found = True
            break
    if not found and jogador:
        nr = ws.max_row + 1
        for col, val in enumerate([jogador, time_gol, comp, 1, 0, 1], 1):
            c = ws.cell(nr, col, val)
            c.border = border
            c.alignment = Alignment(horizontal="center")

    # Busca ou cria linha do assistente
    if assist:
        found2 = False
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == assist and ws.cell(r, 3).value == comp:
                ws.cell(r, 5).value = (ws.cell(r, 5).value or 0) + 1
                found2 = True
                break
        if not found2:
            nr = ws.max_row + 1
            time_assist = row.get("time_mand", "")
            for col, val in enumerate([assist, time_assist, comp, 0, 1, 1], 1):
                c = ws.cell(nr, col, val)
                c.border = border
                c.alignment = Alignment(horizontal="center")

def _atualizar_resumo_time(wb, row):
    """Atualiza gols marcados/sofridos por time."""
    from openpyxl.styles import Alignment, Border, Side
    ws = wb["Resumo por Time"]
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    time_gol = row.get("time_gol", "")
    time_mand = row.get("time_mand", "")
    time_visit = row.get("time_visit", "")
    comp = row.get("competicao", "")
    time_sofreu = time_visit if time_gol == time_mand else time_mand

    for time, marcou in [(time_gol, True), (time_sofreu, False)]:
        if not time:
            continue
        found = False
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == time and ws.cell(r, 2).value == comp:
                if marcou:
                    ws.cell(r, 4).value = (ws.cell(r, 4).value or 0) + 1
                else:
                    ws.cell(r, 5).value = (ws.cell(r, 5).value or 0) + 1
                found = True
                break
        if not found:
            nr = ws.max_row + 1
            gols_m = 1 if marcou else 0
            gols_s = 0 if marcou else 1
            for col, val in enumerate([time, comp, 1, gols_m, gols_s], 1):
                c = ws.cell(nr, col, val)
                c.border = border
                c.alignment = Alignment(horizontal="center")

# Mantém CSV simples como backup extra
def csv_ensure():
    if not CSV_PATH.exists():
        with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow([
                "data","hora","competicao","time_mand","time_visit","time_gol",
                "timestamp_video","jogador_gol","num_gol",
                "jogador_assist","num_assist","video_origem"
            ])

def csv_save(row: dict):
    csv_ensure()
    with open(CSV_PATH, "a", newline="", encoding="utf-8") as f:
        csv.writer(f).writerow([
            row.get("data",""), row.get("hora",""),
            row.get("competicao",""), row.get("time_mand",""),
            row.get("time_visit",""), row.get("time_gol",""),
            row.get("timestamp_video",""),
            row.get("jogador_gol",""), row.get("num_gol",""),
            row.get("jogador_assist",""), row.get("num_assist",""),
            row.get("video_origem",""),
        ])

# ─── VÍDEO ───────────────────────────────────────────────────────────────────
def get_three_frames_b64(video_path: str) -> list:
    import base64
    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS) or 25
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    duration = total_frames / fps
    t1 = 20 * 60
    t2 = duration / 2
    t3 = max(0, duration - 20 * 60)
    timestamps = [
        (t1, "🕐 +20min do início"),
        (t2, f"🕑 Meio ({int(t2//60)}min)"),
        (t3, f"🕒 -20min do final ({int(t3//60)}min)"),
    ]
    results = []
    for t, label in timestamps:
        cap.set(cv2.CAP_PROP_POS_MSEC, t * 1000)
        ret, frame = cap.read()
        if not ret:
            continue
        h, w = frame.shape[:2]
        frame = cv2.resize(frame, (960, int(h * 960 / w)))
        _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
        results.append({"b64": base64.b64encode(buf).decode(), "label": label, "seconds": t})
    cap.release()
    return results

def get_frame_size(video_path: str):
    cap = cv2.VideoCapture(video_path)
    w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    cap.release()
    return w, h

def ts_fmt(seconds: float) -> str:
    s = int(seconds)
    return f"{s//3600:02d}:{(s%3600)//60:02d}:{s%60:02d}"

def _convert_dropbox_url(url):
    url = re.sub(r'[?&]dl=0', '?dl=1', url)
    if 'dl=1' not in url:
        url += '?dl=1' if '?' not in url else '&dl=1'
    return url.replace('www.dropbox.com', 'dl.dropboxusercontent.com')

def _download_direct(url, session_id):
    import requests as req
    out_path = WORK_DIR / f"video_{session_id}.mp4"
    try:
        with req.get(url, stream=True, timeout=600, allow_redirects=True) as r:
            r.raise_for_status()
            with open(out_path, 'wb') as f:
                for chunk in r.iter_content(chunk_size=8192):
                    f.write(chunk)
        if out_path.exists() and out_path.stat().st_size > 1000:
            return str(out_path), out_path.name, None
        return None, None, "Arquivo vazio"
    except Exception as e:
        return None, None, str(e)

def _download_url(url: str, session_id: str):
    base_out = str(WORK_DIR / f"video_{session_id}")

    if 'drive.google.com' in url:
        log.info("Google Drive detectado")
        try:
            import gdown
            out_path = str(WORK_DIR / f"video_{session_id}.mp4")
            gdown.download(url, out_path, quiet=False)
            if os.path.exists(out_path) and os.path.getsize(out_path) > 1000:
                return out_path, f"video_{session_id}.mp4", None
            return None, None, "Arquivo vazio após download"
        except Exception as e:
            return None, None, str(e)

    if 'dropbox.com' in url or 'dropboxusercontent.com' in url:
        log.info("Dropbox detectado")
        return _download_direct(_convert_dropbox_url(url), session_id)

    last_error = ""
    for fmt in ["best", "worst", "bestvideo+bestaudio/best"]:
        ydl_opts = {
            "format": fmt,
            "outtmpl": base_out + ".%(ext)s",
            "quiet": True,
            "no_warnings": True,
            "noplaylist": True,
            "merge_output_format": "mp4",
            "cookiefile": str(WORK_DIR / "cookies.txt") if (WORK_DIR / "cookies.txt").exists() else None,
        }
        try:
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                info = ydl.extract_info(url, download=True)
                video_name = info.get("title", f"video_{session_id}") + ".mp4"
            candidates = list(WORK_DIR.glob(f"video_{session_id}.*"))
            if candidates:
                return str(candidates[0]), video_name, None
        except Exception as e:
            last_error = str(e)
            continue
    return None, None, last_error

# ─── FFMPEG FUNCTIONS ────────────────────────────────────────────────────────
def detectar_segmentos(video_path: str, X, Y, W, H) -> list:
    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS) or 25
    prev = None
    moving = False
    start = 0
    segments = []
    frame_id = 0
    last_change_time = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        crop = frame[Y:Y+H, X:X+W]
        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
        gray = cv2.GaussianBlur(gray, (5, 5), 0)
        tempo = frame_id / fps

        if prev is not None:
            score = np.mean(cv2.absdiff(prev, gray))
            if score > DIFF_THRESHOLD:
                last_change_time = tempo
                if not moving:
                    moving = True
                    start = tempo
            if moving and (tempo - last_change_time > MIN_STATIC):
                end = last_change_time
                if end - start > 2:
                    segments.append((start, end))
                moving = False

        prev = gray
        frame_id += 1

    cap.release()
    return segments

def cortar_segmento(video, start, end, output, padding=1.5):
    subprocess.run([
        FFMPEG, "-y",
        "-ss", str(max(0, start - padding)),
        "-i", video,
        "-t", str((end - start) + padding * 2),
        "-vf", "scale=trunc(iw/2)*2:trunc(ih/2)*2,setsar=1",
        "-c:v", "libx264", "-preset", "fast", "-crf", "28",
        "-c:a", "aac", output
    ], capture_output=True)
    return os.path.exists(output)

def juntar_clips(clips, output):
    if not clips:
        return False
    lista = str(WORK_DIR / "lista.txt")
    with open(lista, "w") as f:
        for c in clips:
            f.write(f"file '{os.path.abspath(c)}'\n")
    subprocess.run([
        FFMPEG, "-y", "-f", "concat", "-safe", "0",
        "-i", lista, "-c", "copy", output
    ], capture_output=True)
    try: os.remove(lista)
    except: pass
    return os.path.exists(output)

def comprimir_video(input_video, output):
    subprocess.run([
        FFMPEG, "-y", "-i", input_video,
        "-vf", "scale=trunc(iw/2)*2:trunc(ih/2)*2,setsar=1",
        "-c:v", "libx264", "-crf", "32", "-preset", "fast",
        "-c:a", "aac", output
    ], capture_output=True)
    return os.path.exists(output)

# ─── ANALISE.IO API ──────────────────────────────────────────────────────────
import requests as _req

def _bot_headers():
    return {"X-Bot-Secret": BOT_SECRET, "Content-Type": "application/json"}

def analise_buscar_atleta(nome: str, clube: str = "") -> list:
    try:
        params = {"nome": nome}
        if clube:
            params["clube"] = clube
        r = _req.get(f"{ANALISE_URL}/api/bot/atletas", params=params, headers=_bot_headers(), timeout=10)
        return r.json() if r.ok else []
    except Exception as e:
        log.error(f"analise_buscar_atleta error: {e}")
        return []

def analise_registrar_gol(atleta_id: str) -> dict:
    try:
        r = _req.post(f"{ANALISE_URL}/api/bot/gol", json={"atleta_id": atleta_id}, headers=_bot_headers(), timeout=10)
        return r.json() if r.ok else {}
    except Exception as e:
        log.error(f"analise_registrar_gol error: {e}")
        return {}

def analise_registrar_assistencia(atleta_id: str) -> dict:
    try:
        r = _req.post(f"{ANALISE_URL}/api/bot/assistencia", json={"atleta_id": atleta_id}, headers=_bot_headers(), timeout=10)
        return r.json() if r.ok else {}
    except Exception as e:
        log.error(f"analise_registrar_assistencia error: {e}")
        return {}

def analise_registrar_jogo(clube: str) -> dict:
    try:
        r = _req.post(f"{ANALISE_URL}/api/bot/jogo", json={"clube": clube}, headers=_bot_headers(), timeout=10)
        return r.json() if r.ok else {}
    except Exception as e:
        log.error(f"analise_registrar_jogo error: {e}")
        return {}

# ─── FLASK ───────────────────────────────────────────────────────────────────
flask_app = Flask(__name__, template_folder="templates")

@flask_app.route("/")
def index():
    return "<h2>⚽ ScoutBot online!</h2><p>Mande um link no Telegram para começar.</p>"

@flask_app.route("/roi/<session_id>")
def roi_page(session_id):
    try:
        s = session_load(session_id)
        if not s:
            return "<h2>⚠️ Sessão expirada</h2><p>Mande o link do Drive novamente no Telegram.</p>", 404
        video_path = s.get("video_path", "")
        if not video_path or not os.path.exists(video_path):
            return "<h2>⚠️ Vídeo não encontrado</h2><p>Mande o link do Drive novamente no Telegram.</p>", 404
        return render_template("select_roi.html", session_id=session_id)
    except Exception as e:
        log.error(f"roi_page error: {e}")
        return f"<h2>Erro</h2><p>{e}</p>", 500

@flask_app.route("/api/frame/<session_id>")
def api_frame(session_id):
    try:
        s = session_load(session_id)
        if not s:
            return jsonify({"ready": False, "error": "sessao_nao_encontrada"})
        if session_id not in _frames_cache:
            video_path = s.get("video_path", "")
            if video_path and os.path.exists(video_path):
                frames = get_three_frames_b64(video_path)
                _frames_cache[session_id] = frames
                _frame_b64_cache[session_id] = frames[0]["b64"] if frames else ""
            else:
                return jsonify({"ready": False, "error": "video_nao_encontrado"})
        frames = _frames_cache.get(session_id, [])
        return jsonify({"frames": frames, "frame": _frame_b64_cache.get(session_id, ""), "ready": bool(frames)})
    except Exception as e:
        log.error(f"api_frame error: {e}")
        return jsonify({"ready": False, "error": str(e)})

@flask_app.route("/api/roi/<session_id>", methods=["POST"])
def api_roi(session_id):
    try:
        s = session_load(session_id)
        if not s:
            return jsonify({"ok": False, "error": "Sessão inválida"})
        data = request.json
        vw, vh = get_frame_size(s["video_path"])
        X = max(0, int(data["x"] * vw))
        Y = max(0, int(data["y"] * vh))
        W = max(1, min(int(data["w"] * vw), vw - X))
        H = max(1, min(int(data["h"] * vh), vh - Y))
        s["roi"] = [X, Y, W, H]
        s["roi_ready"] = True
        session_save(session_id, s)
        log.info(f"ROI salvo {session_id}: {X},{Y},{W},{H}")
        return jsonify({"ok": True, "roi": [X, Y, W, H]})
    except Exception as e:
        log.error(f"api_roi error: {e}")
        return jsonify({"ok": False, "error": str(e)})

@flask_app.route("/api/csv")
def download_csv():
    csv_ensure()
    return send_file(str(CSV_PATH.resolve()), as_attachment=True, download_name="gols.csv")

@flask_app.route("/api/xlsx")
def download_xlsx():
    xlsx_ensure()
    return send_file(str(XLSX_PATH.resolve()), as_attachment=True, download_name="scoutbot_dados.xlsx")

def run_flask():
    flask_app.run(host="0.0.0.0", port=PORT, debug=False, use_reloader=False)

# ─── BOT ─────────────────────────────────────────────────────────────────────
AGUARD_COMPETICAO, AGUARD_COMPETICAO_OUTRO, AGUARD_TIME_MAND, AGUARD_TIME_VISIT, AGUARD_TIME_GOL, AGUARD_NUM_GOL, AGUARD_NOME_GOL, AGUARD_CONFIRMA_ATLETA, AGUARD_ASSIST, AGUARD_NUM_ASSIST, AGUARD_NOME_ASSIST, AGUARD_CONFIRMA_ASSIST, AGUARD_MAIS_GOLS = range(13)
URL_PATTERN = re.compile(r'https?://\S+')

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "⚽ *ScoutBot ativo!*\n\nCola um link (Drive ou Dropbox).\n\n"
        "/csv — planilha\n/status — servidor\n/cancelar — cancelar",
        parse_mode="Markdown"
    )

async def cmd_status(update, ctx):
    cookies_path = WORK_DIR / "cookies.txt"
    exists = cookies_path.exists()
    size = cookies_path.stat().st_size if exists else 0
    files = [f.name for f in WORK_DIR.iterdir()] if WORK_DIR.exists() else []
    await update.message.reply_text(
        "WORK_DIR: " + str(WORK_DIR) + "\ncookies.txt: " + str(exists) +
        "\nTamanho: " + str(size) + " bytes\nArquivos: " + str(files) +
        "\nFFMPEG: " + FFMPEG
    )

async def cmd_csv(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    csv_ensure()
    xlsx_ensure()
    await update.message.reply_document(
        document=open(CSV_PATH, "rb"), filename="gols.csv", caption="📊 Backup CSV"
    )
    await update.message.reply_document(
        document=open(XLSX_PATH, "rb"), filename="scoutbot_dados.xlsx",
        caption="📊 Planilha completa (Excel)"
    )

async def handle_cookies(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc or "cookies" not in doc.file_name.lower():
        return
    tg_file = await ctx.bot.get_file(doc.file_id)
    await tg_file.download_to_drive(str(WORK_DIR / "cookies.txt"))
    await update.message.reply_text("✅ *cookies.txt salvo!*", parse_mode="Markdown")

async def handle_video(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    file_obj = update.message.video or update.message.document
    if not file_obj:
        return
    session_id = str(uuid.uuid4())[:8]
    try:
        await update.message.reply_text("📥 Baixando arquivo... aguarda.")
        tg_file = await ctx.bot.get_file(file_obj.file_id, read_timeout=300, write_timeout=300, connect_timeout=300)
        video_path = str(WORK_DIR / f"video_{session_id}.mp4")
        await tg_file.download_to_drive(video_path, read_timeout=600)
    except Exception as e:
        await update.message.reply_text(f"❌ Erro:\n`{e}`", parse_mode="Markdown")
        return
    video_name = getattr(file_obj, "file_name", None) or f"video_{session_id}.mp4"
    frames = get_three_frames_b64(video_path)
    _frames_cache[session_id] = frames
    _frame_b64_cache[session_id] = frames[0]["b64"] if frames else ""
    session_save(session_id, {"video_path": video_path, "video_name": video_name, "roi": None, "roi_ready": False})
    ctx.user_data["session_id"] = session_id
    roi_url = f"{PUBLIC_URL.rstrip('/')}/roi/{session_id}"
    await update.message.reply_text(
        f"✅ Vídeo recebido!\n\n👇 Marque o placar:\n{roi_url}\n\nDepois mande /pronto",
        disable_web_page_preview=True
    )

async def handle_link(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    match = URL_PATTERN.search(text)
    if not match:
        return
    url = match.group(0)
    session_id = str(uuid.uuid4())[:8]
    ctx.user_data["session_id"] = session_id
    msg = await update.message.reply_text("⏳ Baixando vídeo no Railway... pode levar alguns minutos.")
    loop = asyncio.get_event_loop()
    video_path, video_name, error = await loop.run_in_executor(None, _download_url, url, session_id)
    if error or not video_path:
        await msg.edit_text(f"❌ Erro:\n`{error}`\n\nVerifica se o link é público.", parse_mode="Markdown")
        return
    frames = get_three_frames_b64(video_path)
    _frames_cache[session_id] = frames
    _frame_b64_cache[session_id] = frames[0]["b64"] if frames else ""
    session_save(session_id, {"video_path": video_path, "video_name": video_name, "roi": None, "roi_ready": False})
    roi_url = f"{PUBLIC_URL.rstrip('/')}/roi/{session_id}"
    await msg.edit_text(
        f"✅ *{video_name}*\n\n👇 Marque o placar:\n{roi_url}\n\nDepois mande /pronto",
        parse_mode="Markdown", disable_web_page_preview=True
    )

async def cmd_pronto(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    session_id = ctx.user_data.get("session_id")
    s = session_load(session_id)
    if not s:
        await update.message.reply_text("❌ Nenhum vídeo. Manda um link primeiro.")
        return
    if not s.get("roi_ready"):
        roi_url = f"{PUBLIC_URL.rstrip('/')}/roi/{session_id}"
        await update.message.reply_text(f"⚠️ Área não selecionada!\n\nAcesse: {roi_url}")
        return
    await update.message.reply_text("🔍 Analisando... aguarda os clipes!")
    asyncio.ensure_future(_processar(session_id, ctx))

async def _processar(session_id: str, ctx: ContextTypes.DEFAULT_TYPE):
    s = session_load(session_id)
    if not s:
        await ctx.bot.send_message(CHAT_ID, "❌ Sessão expirada.")
        return
    video_path = s["video_path"]
    video_name = s["video_name"]
    X, Y, W, H = s["roi"]

    if not os.path.exists(video_path):
        await ctx.bot.send_message(CHAT_ID, "❌ Vídeo não encontrado. Manda o link de novo.")
        return

    await ctx.bot.send_message(CHAT_ID, f"⚙️ Detectando mudanças em *{video_name}*...", parse_mode="Markdown")

    segments = detectar_segmentos(video_path, X, Y, W, H)

    if not segments:
        await ctx.bot.send_message(CHAT_ID, "⚠️ Nenhuma mudança detectada. Tenta selecionar uma área maior.")
        try: os.remove(video_path)
        except: pass
        session_delete(session_id)
        return

    await ctx.bot.send_message(CHAT_ID, f"✅ {len(segments)} mudança(s) detectada(s)! Cortando e enviando...")

    clips = []
    for j, (s_time, e_time) in enumerate(segments):
        out = str(WORK_DIR / f"clip_{session_id}_{j}.mp4")
        if cortar_segmento(video_path, s_time, e_time, out):
            clips.append(out)

    if not clips:
        await ctx.bot.send_message(CHAT_ID, "❌ Erro ao cortar clipes.")
        return

    final = str(WORK_DIR / f"final_{session_id}.mp4")
    final_small = str(WORK_DIR / f"final_small_{session_id}.mp4")

    ok = juntar_clips(clips, final)
    if not ok:
        await ctx.bot.send_message(CHAT_ID, "❌ Erro ao juntar clipes.")
        return

    ok = comprimir_video(final, final_small)
    if not ok:
        final_small = final

    ts_inicio = ts_fmt(segments[0][0])
    ts_fim = ts_fmt(segments[-1][1])
    clip_id = f"{session_id}_final"
    pending_clips[clip_id] = {"path": final_small, "timestamp": ts_inicio, "video_origem": video_name}

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("⚽ Registrar gols", callback_data=f"gol:{clip_id}")
    ]])

    try:
        with open(final_small, "rb") as f:
            await ctx.bot.send_document(
                chat_id=CHAT_ID,
                document=f,
                filename=f"gols_{video_name}",
                caption=(
                    f"⚽ *{len(segments)} mudança(s) de placar*\n"
                    f"🕐 {ts_inicio} → {ts_fim}\n"
                    f"Clique para registrar os gols."
                ),
                reply_markup=keyboard,
                parse_mode="Markdown",
            )
    except Exception as e:
        log.error(f"Erro ao enviar documento: {e}")
        await ctx.bot.send_message(CHAT_ID, f"❌ Erro ao enviar: {e}")

    for c in clips:
        try: os.remove(c)
        except: pass
    try: os.remove(final)
    except: pass
    try: os.remove(video_path)
    except: pass
    session_delete(session_id)
    _frames_cache.pop(session_id, None)
    _frame_b64_cache.pop(session_id, None)

# ─── REGISTRO DE GOL ─────────────────────────────────────────────────────────
async def cb_registrar(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    clip_id = query.data.split(":", 1)[1]
    ctx.user_data["clip_id"] = clip_id
    clip = pending_clips.get(clip_id, {})
    ctx.user_data["timestamp"] = clip.get("timestamp", "?")
    ctx.user_data["video_origem"] = clip.get("video_origem", "?")
    kb = ReplyKeyboardMarkup(
        [["LNF", "Paranaense"], ["Gaúcho", "Paulista"], ["Outro"]],
        one_time_keyboard=True, resize_keyboard=True
    )
    await query.message.reply_text("🏆 Qual competição?", reply_markup=kb)
    return AGUARD_COMPETICAO

async def recv_competicao(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    resp = update.message.text.strip()
    if resp == "Outro":
        await update.message.reply_text("Digite o nome da competição:")
        return AGUARD_COMPETICAO_OUTRO
    ctx.user_data["competicao"] = resp
    await update.message.reply_text("🏟 Time *mandante*:", parse_mode="Markdown")
    return AGUARD_TIME_MAND

async def recv_competicao_outro(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["competicao"] = update.message.text.strip()
    await update.message.reply_text("🏟 Time *mandante*:", parse_mode="Markdown")
    return AGUARD_TIME_MAND

async def recv_time_mand(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["time_mand"] = update.message.text.strip()
    await update.message.reply_text("Time **visitante**:", parse_mode="Markdown")
    return AGUARD_TIME_VISIT

async def recv_time_visit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["time_visit"] = update.message.text.strip()
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, analise_registrar_jogo, ctx.user_data["time_mand"])
    await loop.run_in_executor(None, analise_registrar_jogo, ctx.user_data["time_visit"])
    kb = ReplyKeyboardMarkup(
        [[ctx.user_data["time_mand"], ctx.user_data["time_visit"]]],
        one_time_keyboard=True, resize_keyboard=True
    )
    await update.message.reply_text(
        f"✅ *{ctx.user_data['time_mand']}* x *{ctx.user_data['time_visit']}*\n\nDe qual time foi o gol?",
        parse_mode="Markdown", reply_markup=kb
    )
    return AGUARD_TIME_GOL

async def recv_time_gol(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["time_gol"] = update.message.text.strip()
    await update.message.reply_text("Número da camisa do goleador:")
    return AGUARD_NUM_GOL

async def recv_num_gol(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["num_gol"] = update.message.text.strip()
    await update.message.reply_text("Nome do goleador:")
    return AGUARD_NOME_GOL

async def recv_nome_gol(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    nome = update.message.text.strip()
    ctx.user_data["jogador_gol"] = nome
    # Busca no analise.io
    clube = ctx.user_data.get("time_mand", "")
    loop = asyncio.get_event_loop()
    atletas = await loop.run_in_executor(None, analise_buscar_atleta, nome, clube)
    if atletas:
        ctx.user_data["atletas_encontrados"] = atletas
        lista = "\n".join([f"{i+1}. *{a['nome']}* — {a['clube']} ({a['posicao']}) | ⚽{a.get('stats_gols','0')} 🅰️{a.get('stats_assists','0')} 🎮{a.get('stats_jogos','0')}" for i,a in enumerate(atletas[:5])])
        kb = ReplyKeyboardMarkup(
            [[str(i+1) for i in range(min(len(atletas),5))], ["Não encontrado"]],
            one_time_keyboard=True, resize_keyboard=True
        )
        await update.message.reply_text(
            f"Atletas encontrados:\n{lista}\n\nQual é o correto?",
            parse_mode="Markdown", reply_markup=kb
        )
        return AGUARD_CONFIRMA_ATLETA
    else:
        ctx.user_data["atleta_gol_id"] = None
        kb = ReplyKeyboardMarkup([["Sim", "Não"]], one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text("⚠️ Atleta não encontrado no site.\n\nTeve assistência?", reply_markup=kb)
        return AGUARD_ASSIST

async def recv_confirma_atleta(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    resp = update.message.text.strip()
    atletas = ctx.user_data.get("atletas_encontrados", [])
    if resp.isdigit() and 1 <= int(resp) <= len(atletas):
        atleta = atletas[int(resp)-1]
        ctx.user_data["atleta_gol_id"] = atleta["id"]
        ctx.user_data["jogador_gol"] = atleta["nome"]
        await update.message.reply_text(f"✅ *{atleta['nome']}* selecionado!", parse_mode="Markdown")
    else:
        ctx.user_data["atleta_gol_id"] = None
    kb = ReplyKeyboardMarkup([["Sim", "Não"]], one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text("Teve assistência?", reply_markup=kb)
    return AGUARD_ASSIST

async def recv_assist(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.message.text.strip().lower() in ("sim", "s"):
        await update.message.reply_text("Número da camisa de quem assistiu:")
        return AGUARD_NUM_ASSIST
    return await _salvar(update, ctx, "", "")

async def recv_num_assist(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["num_assist"] = update.message.text.strip()
    await update.message.reply_text("Nome de quem assistiu:")
    return AGUARD_NOME_ASSIST

async def recv_nome_assist(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    nome = update.message.text.strip()
    # Busca no analise.io
    clube = ctx.user_data.get("time_mand", "")
    loop = asyncio.get_event_loop()
    atletas = await loop.run_in_executor(None, analise_buscar_atleta, nome, clube)
    if atletas:
        ctx.user_data["atletas_assist_encontrados"] = atletas
        lista = "\n".join([f"{i+1}. *{a['nome']}* — {a['clube']}" for i,a in enumerate(atletas[:5])])
        kb = ReplyKeyboardMarkup(
            [[str(i+1) for i in range(min(len(atletas),5))], ["Não encontrado"]],
            one_time_keyboard=True, resize_keyboard=True
        )
        await update.message.reply_text(
            f"Atletas encontrados:\n{lista}\n\nQual é o correto?",
            parse_mode="Markdown", reply_markup=kb
        )
        ctx.user_data["nome_assist_tmp"] = nome
        return AGUARD_CONFIRMA_ASSIST
    else:
        ctx.user_data["atleta_assist_id"] = None
        return await _salvar(update, ctx, ctx.user_data.get("num_assist", ""), nome)

async def recv_confirma_assist(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    resp = update.message.text.strip()
    atletas = ctx.user_data.get("atletas_assist_encontrados", [])
    if resp.isdigit() and 1 <= int(resp) <= len(atletas):
        atleta = atletas[int(resp)-1]
        ctx.user_data["atleta_assist_id"] = atleta["id"]
        nome_assist = atleta["nome"]
    else:
        ctx.user_data["atleta_assist_id"] = None
        nome_assist = ctx.user_data.get("nome_assist_tmp", "")
    return await _salvar(update, ctx, ctx.user_data.get("num_assist", ""), nome_assist)

async def _salvar(update, ctx, num_assist, nome_assist):
    now = datetime.now()
    row_data = {
        "data": now.strftime("%d/%m/%Y"), "hora": now.strftime("%H:%M:%S"),
        "competicao": ctx.user_data.get("competicao", ""),
        "time_mand": ctx.user_data.get("time_mand", ""),
        "time_visit": ctx.user_data.get("time_visit", ""),
        "time_gol": ctx.user_data.get("time_gol", ""),
        "timestamp_video": ctx.user_data.get("timestamp", ""),
        "jogador_gol": ctx.user_data.get("jogador_gol", ""),
        "num_gol": ctx.user_data.get("num_gol", ""),
        "jogador_assist": nome_assist, "num_assist": num_assist,
        "video_origem": ctx.user_data.get("video_origem", ""),
    }
    csv_save(row_data)
    xlsx_save(row_data)
    clip_id = ctx.user_data.get("clip_id", "")
    if clip_id in pending_clips:
        try: os.remove(pending_clips[clip_id]["path"])
        except: pass
        del pending_clips[clip_id]
    # Atualiza analise.io
    loop = asyncio.get_event_loop()
    atleta_gol_id = ctx.user_data.get("atleta_gol_id")
    atleta_assist_id = ctx.user_data.get("atleta_assist_id")
    site_gol = ""
    site_assist = ""
    if atleta_gol_id:
        res = await loop.run_in_executor(None, analise_registrar_gol, atleta_gol_id)
        if res.get("ok"):
            site_gol = f" _(site: {res.get('gols','?')} gols)_"
    if atleta_assist_id and nome_assist:
        res = await loop.run_in_executor(None, analise_registrar_assistencia, atleta_assist_id)
        if res.get("ok"):
            site_assist = f" _(site: {res.get('assists','?')} assist.)_"

    assist_str = f"\n🅰️ #{num_assist} {nome_assist}{site_assist}" if nome_assist else ""
    jogo_str = f"\n🏟 {ctx.user_data.get('time_mand','')} x {ctx.user_data.get('time_visit','')}" if ctx.user_data.get('time_mand') else ""
    await update.message.reply_text(
        f"✅ *Gol salvo!*\n⚽ #{ctx.user_data.get('num_gol','')} {ctx.user_data.get('jogador_gol','')}{site_gol}{assist_str}{jogo_str}\n"
        f"🕐 {ctx.user_data.get('timestamp','')}\n\n_Salvo no CSV e no site!_ 🌐",
        parse_mode="Markdown"
    )
    kb = ReplyKeyboardMarkup([["Sim", "Não"]], one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text("Tem mais gols nesse jogo para registrar?", reply_markup=kb)
    return AGUARD_MAIS_GOLS

async def recv_mais_gols(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    resp = update.message.text.strip().lower()
    if resp in ("sim", "s"):
        # Volta para perguntar de qual time foi o gol (mantém jogo/competição)
        kb = ReplyKeyboardMarkup(
            [[ctx.user_data.get("time_mand",""), ctx.user_data.get("time_visit","")]],
            one_time_keyboard=True, resize_keyboard=True
        )
        await update.message.reply_text("⚽ De qual time foi o próximo gol?", reply_markup=kb)
        return AGUARD_TIME_GOL
    else:
        # Manda planilha e encerra
        try:
            xlsx_ensure()
            caption = (
                f"📊 *Planilha atualizada!*\n"
                f"🏆 {ctx.user_data.get('competicao','')}\n"
                f"🏟 {ctx.user_data.get('time_mand','')} x {ctx.user_data.get('time_visit','')}"
            )
            with open(XLSX_PATH, "rb") as f:
                await update.message.reply_document(
                    document=f,
                    filename="scoutbot_dados.xlsx",
                    caption=caption,
                    parse_mode="Markdown"
                )
        except Exception as e:
            log.error(f"Erro ao enviar planilha: {e}")
            await update.message.reply_text(f"✅ Jogo encerrado! Use /csv para baixar a planilha.\nErro: {e}")
        return ConversationHandler.END

async def cancelar(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Cancelado.")
    return ConversationHandler.END

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    csv_ensure()
    xlsx_ensure()
    threading.Thread(target=run_flask, daemon=True).start()
    log.info(f"🌐 Flask porta {PORT} | {PUBLIC_URL}")
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(cb_registrar, pattern=r"^gol:")],
        states={
            AGUARD_COMPETICAO:     [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_competicao)],
            AGUARD_COMPETICAO_OUTRO:[MessageHandler(filters.TEXT & ~filters.COMMAND, recv_competicao_outro)],
            AGUARD_TIME_MAND:      [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_time_mand)],
            AGUARD_TIME_VISIT:     [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_time_visit)],
            AGUARD_TIME_GOL:       [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_time_gol)],
            AGUARD_NUM_GOL:        [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_num_gol)],
            AGUARD_NOME_GOL:       [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_nome_gol)],
            AGUARD_CONFIRMA_ATLETA:[MessageHandler(filters.TEXT & ~filters.COMMAND, recv_confirma_atleta)],
            AGUARD_ASSIST:         [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_assist)],
            AGUARD_NUM_ASSIST:     [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_num_assist)],
            AGUARD_NOME_ASSIST:    [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_nome_assist)],
            AGUARD_CONFIRMA_ASSIST:[MessageHandler(filters.TEXT & ~filters.COMMAND, recv_confirma_assist)],
            AGUARD_MAIS_GOLS:      [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_mais_gols)],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)],
        allow_reentry=True,
    )
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("pronto", cmd_pronto))
    app.add_handler(CommandHandler("csv", cmd_csv))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(conv)
    app.add_handler(MessageHandler(filters.VIDEO | filters.Document.VIDEO, handle_video))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_cookies))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_link))
    log.info("🤖 Bot iniciado!")
    app.run_polling()

if __name__ == "__main__":
    main()
